import datetime

from .Utils import Utils
from openpyxl import load_workbook


def get_date_serial(suffix_length=4):
    suffix = Utils().unique_string(suffix_length)
    serial = datetime.datetime.now().strftime('%d%H%M%S')
    return serial + suffix


def read_excel(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active  # Get the first sheet

    excel_data = []
    for row in sheet.iter_rows(values_only=True):  # Read all rows
        excel_data.append([cell if cell is not None else '' for cell in row])  # Handle empty cells

    return excel_data